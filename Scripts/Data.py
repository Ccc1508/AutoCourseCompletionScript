import openpyxl
import pyautogui as p
from typing import List, Dict, Union

# 指定Excel文件的路径
filepath : str = "../Excel/UserInfo.xlsx"

# 加载Excel工作簿，data_only=True表示只读取单元格的实际值，不读取公式
wb : openpyxl.Workbook = openpyxl.load_workbook(filepath, data_only=True)


"""
从Excel表中获取账号信息
这部分代码从Excel文件的"UserInfo"工作表中读取账号信息，并将其存储为二维列表。
"""
UserInfo: List[List[Union[int, str]]] = [
    [int(cell.value) if idx == 0 else cell.value for idx, cell in enumerate(row)]
    for row in wb['UserInfo'].iter_rows()
]


"""
从Excel表中获取讨论信息
这部分代码从Excel文件的"Discuss"工作表中读取讨论信息，存储为字典，键是整数，值是列表。
"""
Discuss : Dict[int, List[str]] = {}
for row in wb['Discuss'].iter_rows(min_row=1):
    key = int(row[0].value)
    if key is not None:
        Discuss[key] = [cell.value for cell in row[1:]]


"""
从Excel表中读取笔记信息
这部分代码从Excel文件的"Notes"工作表中读取笔记信息，存储为一维列表。
"""
Notes: List[str] = [
    row[0].value for row in wb['Notes'].iter_rows()
]

"""
屏幕分辨率
"""
screen_width : int = p.size().width
screen_height : int = p.size().height
